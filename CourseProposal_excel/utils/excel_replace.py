import json
import openpyxl
import re  # For cell reference validation
import os
from helpers import load_json_file
import pandas as pd
from excel_conversion_pipeline import create_course_dataframe

def process_excel_with_direct_mapping(json_data_path, excel_template_path, output_excel_path, ensemble_output_path):
    """
    Processes an Excel template by directly placing JSON values into specified cells
    based on a pre-defined cell_replacement_map.  Assumes JSON already contains the final values.

    Args:
        json_data_path (str): Path to the JSON file containing the data (generated_mapping.json),
                              assumed to have keys with pre-concatenated values.
        excel_template_path (str): Path to the Excel template file (.xlsx).
        output_excel_path (str): Path to save the processed Excel file.
    """

    json_data = load_json_file(json_data_path)
    if not json_data:
        print("Failed to load JSON data. Exiting.")
        return
    
    ensemble_output = load_json_file(ensemble_output_path)
    if not ensemble_output:
        print("Failed to load Ensemble Output JSON data. Exiting.")
        return

    sheet1 = "1 - Course Particulars"
    sheet2 = "2 - Background"
    sheet3 = "3 - Instructional Design"
    sheet4 = "4 - Methodologies"

    cell_replacement_map = {  # Define your cell mapping dictionary DIRECTLY HERE
        # Sheet 1 items
        "#Company":      {"sheet": sheet1, "cell": "C2",  "json_key": "#Company"},        # Example - adjust json_key as needed
        "#CourseTitle":   {"sheet": sheet1, "cell": "C3",  "json_key": "#CourseTitle"},     # Example - adjust json_key as needed
        "#TCS_Code_Skill":      {"sheet": sheet1, "cell": "C10",  "json_key": "#TCS_Code_Skill"},      # Map to existing key

        # Sheet 2 items
        "#Placeholder[0]": {"sheet": sheet2, "cell": "B4", "json_key": "#Placeholder[0]"}, # Map to existing key
        "#Placeholder[1]": {"sheet": sheet2, "cell": "B8", "json_key": "#Placeholder[1]"}, # Map to existing key

        # Sheet 3 items
        "#Sequencing_rationale": {"sheet": sheet3, "cell": "B6", "json_key": "#Sequencing_rationale"}, # Map to existing key
        "#Combined_LO": {"sheet": sheet3, "cell": "B4", "json_key": "#Combined_LO"}, # Map to existing key
        # "#Sequencing_rationale": {"sheet": sheet3, "cell": "B6", "json_key": "#Sequencing_rationale"}, # Map to existing key
        # "#Sequencing_rationale": {"sheet": sheet3, "cell": "B6", "json_key": "#Sequencing_rationale"}, # Map to existing key
        # "#Sequencing_rationale": {"sheet": sheet3, "cell": "B6", "json_key": "#Sequencing_rationale"}, # Map to existing key
        # "#Sequencing_rationale": {"sheet": sheet3, "cell": "B6", "json_key": "#Sequencing_rationale"}, # Map to existing key        
    }

    try:
        workbook = openpyxl.load_workbook(excel_template_path, data_only=False)
    except FileNotFoundError:
        print(f"Error: Excel template file not found at {excel_template_path}")
        return

    # --- DataFrame insertion ---
    df = create_course_dataframe(ensemble_output)
    if not df.empty:  # Only proceed if the DataFrame is not empty
        if sheet3 in workbook.sheetnames:
            sheet = workbook[sheet3]
            start_row = 17  # Row after headers (B17 is the cell, so row 18)
            start_col = 2   # Column B (index 2, as openpyxl is 1-indexed)

            # Write DataFrame to Excel
            for r_idx, row in enumerate(df.values):
                for c_idx, value in enumerate(row):
                    sheet.cell(row=start_row + r_idx, column=start_col + c_idx, value=str(value))
        else:
            print(f"Warning: Sheet '{sheet_name}' not found. DataFrame not inserted.")
    else:
        print("Warning: DataFrame is empty. Nothing to insert into Excel.")


    for new_key_name, mapping_config in cell_replacement_map.items():
        if not isinstance(mapping_config, dict) or 'sheet' not in mapping_config or 'cell' not in mapping_config or 'json_key' not in mapping_config:
            print(f"Warning: Invalid mapping format for key '{new_key_name}'. Expected format: {{'sheet': 'SheetName', 'cell': 'CellRef', 'json_key': 'JsonKey'}}")
            continue

        sheet_name = mapping_config['sheet']
        cell_reference = mapping_config['cell']
        json_key_to_use = mapping_config['json_key'] # Get the JSON key to use for this cell

        if not re.match(r'^[A-Z]+[1-9][0-9]*$', cell_reference): # Basic cell reference validation
            print(f"Warning: Invalid cell reference '{cell_reference}' for key '{new_key_name}'. Skipping.")
            continue

        if sheet_name not in workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook for key '{new_key_name}'. Skipping.")
            continue

        sheet = workbook[sheet_name]
        cell = sheet[cell_reference] # Get the cell object

        # Directly get value from JSON using the specified json_key
        cell_value = json_data.get(json_key_to_use)

        if cell_value is not None:
            if isinstance(cell_value, list): # Handle list values if needed (e.g., concatenate or take first element)
                cell.value = "\n".join(map(str, cell_value)) # Concatenate list items with newlines
            else:
                cell.value = str(cell_value) # Convert to string and set cell value
                if not cell.is_date and cell.data_type == 'f':  # 'f' means formula
                    print(f"Skipping formula cell {cell_reference} to avoid overwriting")
                    continue  # Skip cells containing formulas
        else:
            print(f"Warning: JSON key '{json_key_to_use}' not found in JSON data. Cell '{cell_reference}' will be empty.")


    try:
        workbook.close()
        workbook.save(output_excel_path)
        print(f"Updated Excel file saved to: {output_excel_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")



if __name__ == "__main__":
    generated_mapping_path = os.path.join('..', 'json_output', 'generated_mapping.json') # Path to your data JSON - now assumes pre-processed
    excel_template_path = os.path.join('..', 'templates', 'CP_excel_template.xlsx') # Path to your Excel template
    output_excel_path = os.path.join('..', 'output_docs', 'Course Proposal Template V1_output_direct_value_map.xlsx') # Path for output Excel
    ensemble_output_path = os.path.join('..', 'json_output', 'ensemble_output.json')

    process_excel_with_direct_mapping(generated_mapping_path, excel_template_path, output_excel_path, ensemble_output_path)